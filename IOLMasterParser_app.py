#!/usr/bin/env python3
"""IOLMaster 700 CSV -> Eye-based XLSX converter GUI.

For end users: run the packaged IOLMasterParser.exe, choose a CSV export,
and save the converted XLSX file.

For developers: build a Windows no-install executable with build_windows.bat.
"""

import csv
import io
import sys
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from biometry_ood import load_default_model

APP_NAME = "IOLMaster 700 Biometry Parser"

SOURCE_COLUMNS = [
    "Pat_ID", "Last_Name", "First_Name", "DOB", "Acquisition_Date", "Eye_Side",
    "AL", "AL_SD", "R1", "R2", "A1", "A2", "TR1", "TR2", "TA1", "TA2",
    "ACD", "AQD", "LT", "CCT", "W2W", "P", "Sphere", "Cylinder", "Axis",
]

OOD_OUTPUT_COLUMNS = [
    "Age_at_Biometry", "Mean_K", "Age_Adjusted_ACD", "Age_Adjusted_LT",
    "OOD_Percentile", "OOD_Status", "OOD_Reference_Context",
    "OOD_Largest_Marginal_Deviations", "OOD_Local_Calibration_Effective_N",
    "OOD_Local_Calibration_Max_Percentile", "OOD_Calibration_Warning",
    "OOD_Core_Sensitivity_Percentile", "OOD_Core_Sensitivity_Status",
    "OOD_Model_Selection_Warning",
    "OOD_Age_Stratum", "OOD_Model_Tier", "OOD_Model_Version", "OOD_Distance",
]

OUTPUT_COLUMNS = SOURCE_COLUMNS + OOD_OUTPUT_COLUMNS

NUMERIC_COLUMNS = {
    "AL", "AL_SD", "R1", "R2", "A1", "A2", "TR1", "TR2", "TA1", "TA2",
    "ACD", "AQD", "LT", "CCT", "W2W", "P", "Sphere", "Cylinder", "Axis",
    "Age_at_Biometry", "Mean_K", "Age_Adjusted_ACD", "Age_Adjusted_LT",
    "OOD_Distance", "OOD_Percentile", "OOD_Local_Calibration_Effective_N",
    "OOD_Local_Calibration_Max_Percentile",
    "OOD_Core_Sensitivity_Percentile",
}

TEXT_COLUMNS = {
    "Pat_ID", "Last_Name", "First_Name", "DOB", "Acquisition_Date", "Eye_Side",
    "OOD_Status", "OOD_Reference_Context", "OOD_Largest_Marginal_Deviations",
    "OOD_Calibration_Warning", "OOD_Model_Selection_Warning",
    "OOD_Core_Sensitivity_Status", "OOD_Model_Version", "OOD_Age_Stratum", "OOD_Model_Tier",
}

CSV_ENCODINGS = ("utf-8-sig", "utf-16", "cp949")


def parse_value(column: str, value: str):
    """Parse one CSV value into a spreadsheet-friendly value."""
    value = (value or "").strip()
    if value == "":
        return None
    if column in TEXT_COLUMNS:
        return value
    if column in NUMERIC_COLUMNS:
        try:
            number = float(value)
            if column == "Axis" and number.is_integer():
                return int(number)
            return number
        except ValueError:
            return value
    return value


def read_csv_text(csv_path: Path) -> str:
    """Read the export using common encodings seen on Windows hospital PCs."""
    last_error = None
    for encoding in CSV_ENCODINGS:
        try:
            return csv_path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"CSV 파일 인코딩을 읽을 수 없습니다: {last_error}") from last_error


def find_header_line(lines):
    for idx, line in enumerate(lines):
        columns = [col.strip() for col in line.split(";")]
        if "Pat_ID" in columns and "Acquisition_Date" in columns and "Eye_Side" in columns:
            return idx
    return -1


def read_iolmaster_csv(csv_path: Path):
    """Read ZEISS IOLMaster 700 semicolon-delimited CSV and return rows."""
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV 파일을 찾을 수 없습니다: {csv_path}")
    if not csv_path.is_file():
        raise ValueError(f"CSV 파일이 아닙니다: {csv_path}")

    text = read_csv_text(csv_path)
    lines = text.splitlines()
    header_idx = find_header_line(lines)
    if header_idx < 0:
        raise ValueError("CSV 헤더를 찾을 수 없습니다. ZEISS IOLMaster 700 Export CSV인지 확인하세요.")

    reader = csv.DictReader(io.StringIO("\n".join(lines[header_idx:])), delimiter=";")
    fieldnames = reader.fieldnames or []
    missing = [col for col in SOURCE_COLUMNS if col not in fieldnames]
    if missing:
        raise ValueError(
            "CSV 형식이 예상과 다릅니다. 누락된 컬럼: " + ", ".join(missing)
        )
    model = load_default_model()
    rows = []
    for csv_row in reader:
        source_row = {col: parse_value(col, csv_row.get(col, "")) for col in SOURCE_COLUMNS}
        ood_result = model.score_row(source_row)
        rows.append(
            [source_row[col] for col in SOURCE_COLUMNS]
            + [ood_result[col] for col in OOD_OUTPUT_COLUMNS]
        )
    return rows


def autosize_columns(ws, max_width: int = 28):
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for cell in ws[letter][1: min(ws.max_row, 200)]:  # sample first 200 data rows for speed
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def write_eye_based_xlsx(rows, xlsx_path: Path):
    """Write rows to XLSX in Eye_based format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Eye_based"

    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    ood_header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = ood_header_fill if col_idx > len(SOURCE_COLUMNS) else header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(OUTPUT_COLUMNS)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    autosize_columns(ws)

    # Number formats
    numeric_9_dec = ["AL", "AL_SD", "R1", "R2", "A1", "A2", "TR1", "TR2", "TA1", "TA2", "ACD", "AQD", "LT", "CCT", "W2W", "P"]
    numeric_6_dec = ["Mean_K", "Age_Adjusted_ACD", "Age_Adjusted_LT", "OOD_Distance"]
    numeric_3_dec = [
        "Age_at_Biometry", "OOD_Percentile", "OOD_Local_Calibration_Max_Percentile",
        "OOD_Core_Sensitivity_Percentile",
    ]
    numeric_1_dec = ["OOD_Local_Calibration_Effective_N"]
    numeric_2_dec = ["Sphere", "Cylinder"]
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if ws.max_row < 2:
            continue
        if col_name in numeric_9_dec:
            for cell in ws[f"{letter}2:{letter}{ws.max_row}"]:
                cell[0].number_format = "0.000000000"
        elif col_name in numeric_6_dec:
            for cell in ws[f"{letter}2:{letter}{ws.max_row}"]:
                cell[0].number_format = "0.000000"
        elif col_name in numeric_3_dec:
            for cell in ws[f"{letter}2:{letter}{ws.max_row}"]:
                cell[0].number_format = "0.000"
        elif col_name in numeric_1_dec:
            for cell in ws[f"{letter}2:{letter}{ws.max_row}"]:
                cell[0].number_format = "0.0"
        elif col_name in numeric_2_dec:
            for cell in ws[f"{letter}2:{letter}{ws.max_row}"]:
                cell[0].number_format = "0.00"
        elif col_name == "Pat_ID":
            for cell in ws[f"{letter}2:{letter}{ws.max_row}"]:
                cell[0].number_format = "@"
        elif col_name == "Axis":
            for cell in ws[f"{letter}2:{letter}{ws.max_row}"]:
                cell[0].number_format = "0"

    status_col = get_column_letter(OUTPUT_COLUMNS.index("OOD_Status") + 1)
    status_fills = {
        "Typical anatomy": PatternFill("solid", fgColor="E2F0D9"),
        "Uncommon anatomy": PatternFill("solid", fgColor="FFF2CC"),
        "Rare anatomy": PatternFill("solid", fgColor="FCE4D6"),
        "Not calculated": PatternFill("solid", fgColor="E7E6E6"),
    }
    for cell_tuple in ws[f"{status_col}2:{status_col}{ws.max_row}"]:
        cell = cell_tuple[0]
        if cell.value in status_fills:
            cell.fill = status_fills[cell.value]

    if ws.max_row >= 1:
        ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{ws.max_row}"
        table = Table(displayName="EyeBasedTable", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    wb.save(xlsx_path)


def convert_csv_to_xlsx(csv_path: str, xlsx_path: str) -> int:
    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)
    rows = read_iolmaster_csv(csv_file)
    if xlsx_file.suffix.lower() != ".xlsx":
        xlsx_file = xlsx_file.with_suffix(".xlsx")
    write_eye_based_xlsx(rows, xlsx_file)
    return len(rows)


class ConverterApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_NAME)
        self.root.geometry("520x260")
        self.root.resizable(False, False)

        self.csv_path = tk.StringVar()
        self.status = tk.StringVar(value="CSV 파일을 선택해 주세요.")

        pad = {"padx": 22, "pady": 8}
        title = tk.Label(root, text="IOLMaster 700 CSV → Biometry XLSX", font=("맑은 고딕", 15, "bold"))
        title.pack(pady=(22, 6))

        desc = tk.Label(
            root,
            text="CSV를 변환하고 age-adjusted anatomy OOD score를 계산합니다.",
            font=("맑은 고딕", 10),
        )
        desc.pack()

        frame = tk.Frame(root)
        frame.pack(fill="x", **pad)

        entry = tk.Entry(frame, textvariable=self.csv_path, width=52, state="readonly")
        entry.pack(side="left", expand=True, fill="x")
        tk.Button(frame, text="CSV 선택", command=self.choose_csv, width=12).pack(side="left", padx=(8, 0))

        tk.Button(root, text="변환 시작", command=self.convert, height=2, width=22, font=("맑은 고딕", 10, "bold")).pack(pady=12)
        tk.Label(root, textvariable=self.status, fg="#333333", wraplength=460, justify="center").pack(pady=(4, 0))

        note = tk.Label(root, text="※ 연구용 해부학적 희귀도 지표이며 굴절오차 예측값이 아닙니다.", fg="#666666", font=("맑은 고딕", 9))
        note.pack(side="bottom", pady=(0, 12))

    def choose_csv(self):
        selected = filedialog.askopenfilename(
            title="IOLMaster 700 CSV 파일 선택",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if selected:
            self.csv_path.set(selected)
            self.status.set("CSV 선택 완료. 변환 시작을 누르세요.")

    def convert(self):
        if not self.csv_path.get():
            messagebox.showwarning(APP_NAME, "먼저 CSV 파일을 선택해 주세요.")
            return

        input_path = Path(self.csv_path.get())
        default_name = input_path.with_suffix("").name + "_eye_based.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="저장할 XLSX 파일 이름 선택",
            initialdir=str(input_path.parent),
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not output_path:
            self.status.set("저장이 취소되었습니다.")
            return

        try:
            count = convert_csv_to_xlsx(str(input_path), output_path)
            self.status.set(f"변환 완료: {count} eye rows")
            messagebox.showinfo(APP_NAME, f"변환 완료\n\n{count} eye rows 저장됨\n{output_path}")
        except Exception as exc:
            log_path = Path.home() / "IOLMasterParser_error.txt"
            log_path.write_text(traceback.format_exc(), encoding="utf-8")
            self.status.set("오류가 발생했습니다. CSV 형식 또는 파일 열림 상태를 확인하세요.")
            messagebox.showerror(
                APP_NAME,
                f"변환 실패\n\n{exc}\n\n자세한 오류 기록:\n{log_path}",
            )


def main():
    root = tk.Tk()
    try:
        # Makes the window crisp on high-DPI Windows screens when possible.
        if sys.platform.startswith("win"):
            import ctypes
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    ConverterApp(root)
    root.mainloop()


if __name__ == "__main__":
    # Command-line fallback for testing or batch use:
    # IOLMasterParser.exe input.csv output.xlsx
    if len(sys.argv) == 3:
        rows = convert_csv_to_xlsx(sys.argv[1], sys.argv[2])
        print(f"Converted {rows} eye rows to {sys.argv[2]}")
    else:
        main()
