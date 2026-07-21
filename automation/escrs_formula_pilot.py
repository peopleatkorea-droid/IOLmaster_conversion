#!/usr/bin/env python3
"""Prepare and run a deliberately slow ESCRS public-UI research pilot.

The live runner:

* sends only anonymized study identifiers and biometry values;
* operates one visible browser tab with no parallel requests;
* enforces a 45-75 second interval and a daily cap;
* stops on CAPTCHA, HTTP 403/429, or unexpected result structure;
* writes a checkpoint after every case and never overwrites the source workbook.

The ESCRS UI profile must be inspected and explicitly marked verified before the
live command can run.  No internal API calls or endpoint reverse engineering are
used.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import re
import secrets
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import load_workbook

from automation.escrs_pilot_core import (
    FORMULA_ALIASES,
    FormulaCandidate,
    PreparedCase,
    canonical_text,
    finite_float,
    gender_lookup_key,
    parse_formula_segment,
    prepare_case,
    select_evenly_across_al,
    summarize_formula_results,
)


DEFAULT_URL = "https://iolcalculator.escrs.org/"
DEFAULT_KEY_PATH = Path(".escrs_private/escrs_id_key.bin")
DEFAULT_OUTPUT_DIR = Path("analysis_outputs/escrs")
LIVE_CONFIRMATION = "I_UNDERSTAND_ESCRS_UI_AUTOMATION"
BLOCK_TEXT_PATTERNS = (
    "captcha",
    "recaptcha",
    "too many requests",
    "access denied",
    "temporarily blocked",
    "unusual traffic",
    "forbidden",
)


def json_default(value):
    if isinstance(value, (datetime,)):
        return value.isoformat()
    raise TypeError(f"Object is not JSON serializable: {type(value).__name__}")


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=json_default),
        encoding="utf-8",
    )
    temporary.replace(path)


def append_jsonl(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(
            json.dumps(payload, ensure_ascii=False, default=json_default) + "\n"
        )
        handle.flush()
        os.fsync(handle.fileno())


def load_or_create_key(path: Path) -> bytes:
    if path.exists():
        secret = path.read_bytes()
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        secret = secrets.token_bytes(32)
        path.write_bytes(secret)
    if len(secret) < 16:
        raise ValueError("The anonymization key is too short.")
    return secret


def choose_sheet(workbook, requested: str | None, required_headers: set[str]):
    if requested:
        return workbook[requested]
    for worksheet in workbook.worksheets:
        headers = {
            str(cell.value).strip()
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
            if cell.value is not None
        }
        if required_headers <= headers:
            return worksheet
    raise ValueError(
        "No worksheet contains the required headers: "
        + ", ".join(sorted(required_headers))
    )


def worksheet_records(worksheet) -> Iterable[tuple[int, dict[str, object]]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    for excel_row, values in enumerate(rows, start=2):
        yield excel_row, dict(zip(headers, values))


def load_gender_lookup(path: Path, sheet_name: str | None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = choose_sheet(
        workbook,
        sheet_name,
        {"Pat_ID", "Acquisition_Date", "Eye_Side", "Gender"},
    )
    lookup = {}
    conflicts = 0
    for _, row in worksheet_records(worksheet):
        key = gender_lookup_key(row)
        gender = row.get("Gender")
        if not all(key) or gender in (None, ""):
            continue
        if key in lookup and str(lookup[key]).strip() != str(gender).strip():
            conflicts += 1
            continue
        lookup[key] = gender
    workbook.close()
    if conflicts:
        raise ValueError(f"Gender lookup contains {conflicts} conflicting keys.")
    return lookup


def prior_lvc_status(row: Mapping[str, object]) -> str:
    value = str(row.get("Prior_Refractive_Surgery_YN") or "").strip().lower()
    if value in {"y", "yes", "1", "true", "예"}:
        return "yes"
    if value in {"n", "no", "0", "false", "아니오"}:
        return "no"
    return "unverified"


def skip_category(error: ValueError) -> str:
    message = str(error)
    if message.startswith("Missing required values:"):
        return message
    if message.startswith("LT is outside"):
        return "invalid_or_pseudophakic_lens_thickness"
    if message.startswith(("K1 is outside", "K2 is outside")):
        return "keratometry_outside_supported_pilot_range"
    if message.startswith("Gender is missing"):
        return "gender_not_matched"
    return message


def select_pilot_cases(
    prepared: list[tuple[PreparedCase, str, str]],
    limit: int,
) -> tuple[list[PreparedCase], str]:
    cohorts: dict[str, list[PreparedCase]] = defaultdict(list)
    for case, _, cohort in prepared:
        cohorts[cohort or "All"].append(case)
    if len(cohorts) <= 1:
        return (
            select_evenly_across_al([case for case, _, _ in prepared], limit),
            "evenly_across_axial_length",
        )

    names = sorted(cohorts)
    quotas = {name: limit // len(names) for name in names}
    for name in names[: limit % len(names)]:
        quotas[name] += 1
    selected = []
    for name in names:
        selected.extend(select_evenly_across_al(cohorts[name], quotas[name]))
    return (
        sorted(selected, key=lambda case: (case.al_mm, case.public_id)),
        "balanced_by_cohort_then_evenly_across_axial_length",
    )


def prepare_queue(args) -> int:
    input_path = Path(args.input)
    gender_path = Path(args.gender_workbook)
    output_path = Path(args.output)
    secret = load_or_create_key(Path(args.id_key))
    gender_lookup = load_gender_lookup(gender_path, args.gender_sheet)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = choose_sheet(
        workbook,
        args.sheet,
        {"Pat_ID", "Acquisition_Date", "Eye_Side", "AL", "ACD", "LT", "CCT"},
    )
    prepared_by_id: dict[str, tuple[PreparedCase, str, float, str]] = {}
    skipped = Counter()
    for excel_row, row in worksheet_records(worksheet):
        history = prior_lvc_status(row)
        if history == "yes":
            skipped["prior_refractive_surgery_yes"] += 1
            continue
        gender = gender_lookup.get(gender_lookup_key(row))
        try:
            case = prepare_case(row, excel_row, gender, secret)
        except ValueError as exc:
            skipped[skip_category(exc)] += 1
            continue
        al_sd = finite_float(row.get("AL_SD"))
        quality = al_sd if al_sd is not None and al_sd >= 0 else float("inf")
        cohort = canonical_text(row.get("Cohort")) or "All"
        previous = prepared_by_id.get(case.public_id)
        if previous is None or (quality, excel_row) < (previous[2], previous[0].source_row):
            if previous is not None:
                skipped["duplicate_measurement_not_selected"] += 1
            prepared_by_id[case.public_id] = (case, history, quality, cohort)
        else:
            skipped["duplicate_measurement_not_selected"] += 1
    prepared = [
        (case, history, cohort)
        for case, history, _, cohort in prepared_by_id.values()
    ]
    workbook.close()

    selected_cases, selection_method = select_pilot_cases(prepared, args.limit)
    history_by_id = {case.public_id: history for case, history, _ in prepared}
    cohort_by_id = {case.public_id: cohort for case, _, cohort in prepared}
    queue = {
        "schema_version": 1,
        "created_at": datetime.now().astimezone().isoformat(),
        "source_file_name": input_path.name,
        "source_sheet": worksheet.title,
        "selection": selection_method,
        "requested_limit": args.limit,
        "prepared_total": len(prepared),
        "selected_total": len(selected_cases),
        "skipped_counts": dict(skipped),
        "contains_patient_identifiers": False,
        "cases": [
            {
                "case": {
                    **case.browser_payload(),
                    "source_row": case.source_row,
                    "source_eye": case.source_eye,
                },
                "research_flags": {
                    "prior_refractive_surgery_status": history_by_id[case.public_id],
                    "source_cohort": cohort_by_id[case.public_id],
                },
            }
            for case in selected_cases
        ],
    }
    write_json(output_path, queue)
    print(f"Prepared {len(selected_cases)} anonymized pilot cases: {output_path}")
    print(f"Eligible before pilot sampling: {len(prepared)}")
    print(f"Skipped: {sum(skipped.values())}")
    print(
        "Prior refractive surgery remains unverified for "
        f"{sum(history_by_id[case.public_id] == 'unverified' for case in selected_cases)} "
        "selected cases."
    )
    return 0


def locator_manifest(page):
    return page.locator("input, select, textarea, button, [role=combobox]").evaluate_all(
        """elements => elements.map((element, index) => ({
            index,
            tag: element.tagName,
            type: element.getAttribute('type'),
            id: element.id || null,
            name: element.getAttribute('name'),
            role: element.getAttribute('role'),
            ariaLabel: element.getAttribute('aria-label'),
            placeholder: element.getAttribute('placeholder'),
            className: element.className || null,
            readOnly: !!element.readOnly,
            disabled: !!element.disabled,
            text: (element.innerText || element.value || '').trim().slice(0, 200),
            inputControlText: (
                element.closest('.mud-input-control')?.innerText || ''
            ).trim().slice(0, 500),
            ancestors: (() => {
                const result = [];
                let current = element.parentElement;
                for (let depth = 0; current && depth < 5; depth += 1) {
                    result.push({
                        tag: current.tagName,
                        className: current.className || null,
                        text: (current.innerText || '').trim().slice(0, 500)
                    });
                    current = current.parentElement;
                }
                return result;
            })(),
            outerHTML: element.outerHTML.slice(0, 1000),
            visible: !!(element.offsetWidth || element.offsetHeight || element.getClientRects().length)
        }))"""
    )


def visible_text_input(page, index: int):
    locator = page.locator('input[type="text"]:visible')
    if index < 0 or index >= locator.count():
        raise RuntimeError(
            f"Visible text input index {index} is unavailable; "
            f"the page currently has {locator.count()} visible text inputs."
        )
    return locator.nth(index)


def input_control_context(control) -> str:
    return control.evaluate(
        """element => (
            element.closest('.mud-input-control')?.innerText
            || element.parentElement?.innerText
            || ''
        ).trim()"""
    )


def normalized_ui_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def visible_text_input_by_context(
    page,
    patterns: Iterable[str],
    occurrence: int = 0,
):
    matches = []
    locator = page.locator('input[type="text"]:visible')
    for index in range(locator.count()):
        candidate = locator.nth(index)
        context_text = normalized_ui_text(input_control_context(candidate))
        if any(re.search(pattern, context_text, re.I) for pattern in patterns):
            matches.append(candidate)
    if occurrence < 0 or occurrence >= len(matches):
        raise RuntimeError(
            "Visible text input was not found for context patterns "
            f"{list(patterns)!r} at occurrence {occurrence}; "
            f"{len(matches)} matching controls were found."
        )
    return matches[occurrence]


def control_snapshot(page):
    result = []
    locator = page.locator('input[type="text"]:visible')
    for index in range(locator.count()):
        control = locator.nth(index)
        result.append(
            {
                "visible_text_input_index": index,
                "context": normalized_ui_text(input_control_context(control)),
                "value": control.input_value(),
                "read_only": control.is_editable() is False,
                "disabled": control.is_disabled(),
            }
        )
    return result


def visible_option_locators(page):
    selectors = (
        '[role="option"]:visible',
        ".mud-list-item:visible",
        ".mud-popover-open .mud-list-item:visible",
    )
    result = []
    seen = set()
    for selector in selectors:
        locator = page.locator(selector)
        for index in range(locator.count()):
            candidate = locator.nth(index)
            try:
                text = re.sub(r"\s+", " ", candidate.inner_text()).strip()
            except Exception:
                continue
            if not text:
                continue
            if text not in seen:
                seen.add(text)
                result.append((candidate, text))
    return result


def choose_visible_option(page, patterns: Iterable[str]):
    candidates = visible_option_locators(page)
    for pattern in patterns:
        regex = re.compile(pattern, re.I)
        for candidate, text in candidates:
            if regex.fullmatch(text):
                candidate.click()
                return text
    for pattern in patterns:
        regex = re.compile(pattern, re.I)
        for candidate, text in candidates:
            if regex.search(text):
                candidate.click()
                return text
    return None


def probe_dropdowns(page, profile, output_dir: Path, save_progress):
    results = []
    for probe in profile.get("ui_probes", ()):
        name = probe["name"]
        result = {
            "name": name,
            "text_input_index": probe["text_input_index"],
            "context_patterns": probe.get("context_patterns", ()),
            "search_text": probe.get("search_text"),
            "options": [],
            "selected": None,
            "control_value_after_probe": None,
            "controls_after_probe": [],
            "error": None,
        }
        if probe.get("requires_successful_probe"):
            prerequisite = next(
                (
                    item
                    for item in results
                    if item["name"] == probe["requires_successful_probe"]
                ),
                None,
            )
            if prerequisite is None or prerequisite.get("error"):
                result["error"] = (
                    "Skipped because prerequisite probe did not succeed: "
                    + probe["requires_successful_probe"]
                )
                results.append(result)
                save_progress(results)
                continue
        try:
            if probe.get("context_patterns"):
                control = visible_text_input_by_context(
                    page,
                    probe["context_patterns"],
                    int(probe.get("context_occurrence", 0)),
                )
            else:
                control = visible_text_input(
                    page,
                    int(probe["text_input_index"]),
                )
            if control.is_disabled():
                raise RuntimeError(f"Control is disabled for probe: {name}")
            control.click()
            page.wait_for_timeout(700)
            search_text = probe.get("search_text")
            if search_text:
                try:
                    control.fill(str(search_text))
                except Exception:
                    page.keyboard.type(str(search_text), delay=80)
                page.wait_for_timeout(1000)
            result["options"] = [
                text for _, text in visible_option_locators(page)
            ]
            page.screenshot(
                path=str(output_dir / f"probe_{name}_options.png"),
                full_page=True,
            )
            if probe.get("select_option_patterns"):
                result["selected"] = choose_visible_option(
                    page,
                    probe["select_option_patterns"],
                )
                if result["selected"] is None:
                    raise RuntimeError(
                        f"No configured option matched for {name}. "
                        f"Visible options: {result['options'][:30]}"
                    )
                page.wait_for_timeout(1000)
            else:
                page.keyboard.press("Escape")
                page.wait_for_timeout(300)
            try:
                if probe.get("context_patterns"):
                    current_control = visible_text_input_by_context(
                        page,
                        probe["context_patterns"],
                        int(probe.get("context_occurrence", 0)),
                    )
                else:
                    current_control = visible_text_input(
                        page,
                        int(probe["text_input_index"]),
                    )
                result["control_value_after_probe"] = current_control.input_value()
            except Exception:
                result["control_value_after_probe"] = None
            result["controls_after_probe"] = control_snapshot(page)
        except Exception as exc:
            result["error"] = str(exc)
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(300)
            except Exception:
                pass
        finally:
            page.screenshot(
                path=str(output_dir / f"probe_{name}.png"),
                full_page=True,
            )
            results.append(result)
            save_progress(results)
    return results


def calculator_form_is_ready(page, profile) -> bool:
    surgeon = profile.get("fields", {}).get("surgeon", {})
    patterns = surgeon.get("context_patterns", (r"^Surgeon$",))
    try:
        visible_text_input_by_context(page, patterns, 0)
        return True
    except Exception:
        return False


def interactive_control_text(control) -> str:
    return normalized_ui_text(
        control.evaluate(
            """element => (
                element.innerText
                || element.value
                || element.getAttribute('aria-label')
                || element.getAttribute('title')
                || ''
            )"""
        )
    )


def click_visible_terms_control(page, profile) -> str | None:
    labels = [
        normalized_ui_text(label)
        for label in profile.get(
            "terms_accept_names",
            ("I Agree", "Accept", "Agree"),
        )
    ]
    selectors = (
        "button:visible",
        'input[type="button"]:visible',
        'input[type="submit"]:visible',
        '[role="button"]:visible',
        "a:visible",
    )
    for frame in page.frames:
        candidates = []
        seen = set()
        for selector in selectors:
            locator = frame.locator(selector)
            for index in range(locator.count()):
                candidate = locator.nth(index)
                try:
                    text = interactive_control_text(candidate)
                except Exception:
                    continue
                if not text or text in seen:
                    continue
                seen.add(text)
                candidates.append((candidate, text))
        for label in labels:
            for candidate, text in candidates:
                if text.casefold() == label.casefold():
                    candidate.click(timeout=3000)
                    return text
        for label in labels:
            for candidate, text in candidates:
                if label.casefold() in text.casefold():
                    candidate.click(timeout=3000)
                    return text
    return None


def ensure_calculator_ready(page, profile, timeout_seconds: float) -> None:
    deadline = time.time() + timeout_seconds
    clicked_labels = []
    while time.time() < deadline:
        clicked = click_visible_terms_control(page, profile)
        if clicked:
            clicked_labels.append(clicked)
            page.wait_for_timeout(2000)
            continue
        if calculator_form_is_ready(page, profile):
            return
        page.wait_for_timeout(500)
    body_excerpt = ""
    try:
        body_excerpt = normalized_ui_text(page.locator("body").inner_text())[:1000]
    except Exception:
        pass
    raise RuntimeError(
        "The ESCRS calculator form did not become ready after "
        f"{timeout_seconds:.0f}s. Consent controls clicked: {clicked_labels!r}. "
        f"Visible page text: {body_excerpt!r}"
    )


def accept_visible_terms(page, profile) -> None:
    """Backward-compatible one-shot consent helper used by inspect-ui."""
    click_visible_terms_control(page, profile)
    page.wait_for_timeout(1500)


def inspect_ui(args) -> int:
    from playwright.sync_api import sync_playwright

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    profile = read_json(Path(args.profile))
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            channel=args.browser_channel,
            headless=False,
            slow_mo=100,
        )
        page = browser.new_page()
        page.goto(profile.get("url", DEFAULT_URL), wait_until="domcontentloaded")
        page.wait_for_timeout(args.settle_seconds * 1000)
        ensure_calculator_ready(
            page,
            profile,
            args.consent_wait_seconds,
        )
        manifest = {
            "captured_at": datetime.now().astimezone().isoformat(),
            "url": page.url,
            "title": page.title(),
            "controls": locator_manifest(page),
            "body_text_excerpt": page.locator("body").inner_text()[:20000],
        }
        write_json(output_dir / "ui_manifest.json", manifest)
        if args.probe_lens_options:
            def save_probe_progress(results):
                manifest["dropdown_probes"] = list(results)
                write_json(output_dir / "ui_manifest.json", manifest)

            manifest["dropdown_probes"] = probe_dropdowns(
                page,
                profile,
                output_dir,
                save_probe_progress,
            )
        write_json(output_dir / "ui_manifest.json", manifest)
        page.screenshot(path=str(output_dir / "ui_manifest.png"), full_page=True)
        print(f"Saved UI manifest under {output_dir}")
        print("No patient or biometry data were entered.")
        if args.keep_open:
            input("Review the visible browser, then press Enter here to close it...")
        browser.close()
    return 0


def _visible_first(locator):
    for index in range(locator.count()):
        candidate = locator.nth(index)
        if candidate.is_visible():
            return candidate
    return None


def find_control(page, field_name: str, profile):
    field = profile["fields"][field_name]
    if field.get("context_patterns"):
        return visible_text_input_by_context(
            page,
            field["context_patterns"],
            int(field.get("context_occurrence", 0)),
        )
    if field.get("text_input_index") is not None:
        locator = visible_text_input(page, int(field["text_input_index"]))
        context_text = input_control_context(locator)
        patterns = field.get("context_patterns", ())
        if patterns and not any(
            re.search(pattern, context_text, re.I) for pattern in patterns
        ):
            raise RuntimeError(
                f"Control context mismatch for {field_name}: {context_text!r}"
            )
        return locator
    if field.get("css"):
        locator = _visible_first(page.locator(field["css"]))
        if locator is not None:
            return locator
    for label in field.get("labels", ()):
        locator = _visible_first(
            page.get_by_label(re.compile(rf"^{re.escape(label)}", re.I))
        )
        if locator is not None:
            return locator
    for placeholder in field.get("placeholders", ()):
        locator = _visible_first(
            page.get_by_placeholder(re.compile(re.escape(placeholder), re.I))
        )
        if locator is not None:
            return locator
    raise RuntimeError(f"Visible control was not found for field: {field_name}")


def set_control(page, field_name: str, value, profile) -> None:
    field = profile["fields"][field_name]
    control = find_control(page, field_name, profile)
    tag = control.evaluate("element => element.tagName.toLowerCase()")
    role = control.get_attribute("role") or ""
    input_type = (control.get_attribute("type") or "").lower()
    text = str(value)
    if field.get("action") == "mud_dropdown":
        control.click()
        page.wait_for_timeout(400)
        search_text = field.get("search_text")
        if search_text:
            try:
                control.fill(str(search_text))
            except Exception:
                pass
            page.wait_for_timeout(700)
        aliases = field.get("option_patterns_by_value", {}).get(text, ())
        patterns = list(aliases) + list(field.get("option_patterns", ()))
        if not patterns:
            patterns = [rf"^{re.escape(text)}$"]
        selected = choose_visible_option(page, patterns)
        if selected is None:
            visible = [option_text for _, option_text in visible_option_locators(page)]
            raise RuntimeError(
                f"Option not found for {field_name}: {text}. "
                f"Visible options: {visible[:20]}"
            )
        page.wait_for_timeout(1000)
        return
    if tag == "select":
        try:
            control.select_option(label=text)
        except Exception:
            control.select_option(value=text)
        return
    if input_type in {"radio", "checkbox"}:
        control.check()
        return
    if role == "combobox":
        control.click()
        option = _visible_first(
            page.get_by_role("option", name=re.compile(re.escape(text), re.I))
        )
        if option is None:
            option = _visible_first(page.get_by_text(text, exact=False))
        if option is None:
            raise RuntimeError(f"Option not found for {field_name}: {text}")
        option.click()
        return
    control.fill(text)


def validate_page_profile(page, profile) -> None:
    controls = []
    for field_name in profile["field_order"]:
        control = find_control(page, field_name, profile)
        signature = control.evaluate(
            "element => `${element.tagName}|${element.type}|${element.id}`"
        )
        if signature in controls:
            raise RuntimeError(
                f"UI profile maps more than one field to the same control: {field_name}"
            )
        controls.append(signature)
    required_checked = int(profile.get("required_checked_checkbox_count", 0))
    if required_checked:
        checkboxes = page.locator('input[type="checkbox"]:visible')
        if checkboxes.count() < required_checked:
            raise RuntimeError(
                "The expected formula checkboxes are not present on the page."
            )
        unchecked = [
            index
            for index in range(required_checked)
            if not checkboxes.nth(index).is_checked()
        ]
        if unchecked:
            raise RuntimeError(
                "One or more required formula calculators are unchecked: "
                + ", ".join(str(index) for index in unchecked)
            )


def click_named_button(page, names: Iterable[str]):
    for name in names:
        locator = _visible_first(
            page.get_by_role("button", name=re.compile(re.escape(name), re.I))
        )
        if locator is not None:
            locator.click()
            return
    raise RuntimeError("No configured calculate/reset button was visible.")


def fill_case(page, case: Mapping[str, object], profile) -> None:
    static_values = profile.get("static_values", {})
    values = {
        "surgeon": static_values.get("surgeon", "Research"),
        "patient_name": static_values.get("patient_name", "Anonymized"),
        "patient_id": case["public_id"],
        "gender": case["gender"],
        "axial_length": case["al_mm"],
        "k1": case["k1_d"],
        "k2": case["k2_d"],
        "acd": case["acd_mm"],
        "lens_thickness": case["lt_mm"],
        "cct": case["cct_um"],
        "wtw": case["wtw_mm"],
        "target_refraction": case["target_refraction_d"],
        "k_index": case["k_index"],
        "manufacturer": case["manufacturer"],
        "iol_model": case["iol_model"],
    }
    for field_name in profile["field_order"]:
        set_control(page, field_name, values[field_name], profile)


def observed_iol_constants(page, expected):
    controls = page.locator('input[type="text"]')
    observed = {}
    contexts = []
    for index in range(controls.count()):
        control = controls.nth(index)
        context = normalized_ui_text(input_control_context(control))
        try:
            value = control.input_value().strip()
        except Exception:
            continue
        if context or value:
            contexts.append({"context": context, "value": value})
        for label in expected:
            if context.casefold() == normalized_ui_text(label).casefold():
                observed[label] = value
    return observed, contexts


def validate_iol_constants(page, profile, timeout_seconds: float = 15.0) -> None:
    expected = profile.get("expected_iol_constants", {})
    deadline = time.time() + timeout_seconds
    latest_observed = {}
    latest_contexts = []
    while time.time() < deadline:
        latest_observed, latest_contexts = observed_iol_constants(page, expected)
        mismatches = {}
        for label, expected_value in expected.items():
            actual = latest_observed.get(label)
            actual_number = finite_float(actual)
            expected_number = finite_float(expected_value)
            if (
                actual_number is None
                or expected_number is None
                or abs(actual_number - expected_number) > 0.001
            ):
                mismatches[label] = {
                    "expected": str(expected_value),
                    "observed": actual,
                }
        if not mismatches:
            selected_iols = [
                item["value"]
                for item in latest_contexts
                if item["context"].casefold() == "select iol"
                and item["value"]
            ]
            if "ZCB00" not in selected_iols or "ZCB00V" in selected_iols:
                raise RuntimeError(
                    f"Unexpected selected IOL values: {selected_iols!r}"
                )
            return
        page.wait_for_timeout(500)
    available = [
        item
        for item in latest_contexts
        if item["context"]
        and (
            "constant" in item["context"].casefold()
            or "pacd" in item["context"].casefold()
            or item["context"].casefold() == "select iol"
        )
    ]
    raise RuntimeError(
        "ZCB00 constants did not reach the verified values within "
        f"{timeout_seconds:.0f}s. Observed: {latest_observed!r}. "
        f"Relevant controls: {available!r}"
    )


def save_error_snapshot(
    page,
    snapshot_dir: Path,
    public_id: str,
    reason: str,
    response_statuses: Iterable[int],
) -> None:
    base = snapshot_dir / f"{public_id}_error"
    try:
        page.screenshot(path=str(base.with_suffix(".png")), full_page=True)
    except Exception:
        pass
    try:
        base.with_suffix(".txt").write_text(
            page.locator("body").inner_text(),
            encoding="utf-8",
        )
    except Exception:
        pass
    try:
        write_json(
            base.with_suffix(".json"),
            {
                "captured_at": datetime.now().astimezone().isoformat(),
                "url": page.url,
                "reason": reason,
                "response_statuses": list(response_statuses),
                "controls": control_snapshot(page),
            },
        )
    except Exception:
        pass


def split_formula_segments(body_text: str):
    lowered = body_text.lower()
    positions = []
    for formula, aliases in FORMULA_ALIASES.items():
        matches = [
            lowered.find(alias.lower())
            for alias in aliases
            if lowered.find(alias.lower()) >= 0
        ]
        if matches:
            positions.append((min(matches), formula))
    positions.sort()
    segments = {}
    for index, (start, formula) in enumerate(positions):
        end = positions[index + 1][0] if index + 1 < len(positions) else len(body_text)
        segments[formula] = body_text[start:end]
    return segments


def extract_formula_results(page, profile):
    body_text = page.locator("body").inner_text()
    results = {}
    for formula, segment in split_formula_segments(body_text).items():
        candidates = parse_formula_segment(
            segment,
            profile["result_pair_pattern"],
        )
        if candidates:
            results[formula] = candidates
    return body_text, results


def wait_for_formula_results(page, profile, timeout_seconds: int):
    deadline = time.time() + timeout_seconds
    previous_signature = None
    stable_polls = 0
    latest = ("", {})
    while time.time() < deadline:
        latest = extract_formula_results(page, profile)
        body_text, formula_results = latest
        signature = tuple(
            sorted((formula, len(candidates)) for formula, candidates in formula_results.items())
        )
        if len(formula_results) >= profile.get("minimum_result_formulas", 3):
            if signature == previous_signature:
                stable_polls += 1
            else:
                stable_polls = 0
            if stable_polls >= 3:
                return latest
        previous_signature = signature
        page.wait_for_timeout(2000)
    return latest


def completed_ids(checkpoint_path: Path) -> set[str]:
    if not checkpoint_path.exists():
        return set()
    result = set()
    for line in checkpoint_path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        row = json.loads(line)
        if row.get("status") in {"complete", "not_evaluable"}:
            result.add(row["public_id"])
    return result


def interruptible_sleep(seconds: float, reason: str) -> None:
    remaining = max(0.0, seconds)
    while remaining > 0:
        chunk = min(15.0, remaining)
        print(f"{reason}: {remaining:.0f}s remaining", flush=True)
        time.sleep(chunk)
        remaining -= chunk


class RateLimiter:
    def __init__(
        self,
        state_path: Path,
        minimum_interval: float,
        maximum_interval: float,
        daily_limit: int,
        batch_size: int,
        batch_break_seconds: float,
    ):
        self.state_path = state_path
        self.minimum_interval = minimum_interval
        self.maximum_interval = maximum_interval
        self.daily_limit = daily_limit
        self.batch_size = batch_size
        self.batch_break_seconds = batch_break_seconds
        self.state = self._load()

    def _load(self):
        today = datetime.now().astimezone().date().isoformat()
        if self.state_path.exists():
            state = read_json(self.state_path)
        else:
            state = {}
        if state.get("local_date") != today:
            state = {"local_date": today, "daily_count": 0, "last_request_epoch": None}
        return state

    def save(self):
        write_json(self.state_path, self.state)

    def before_case(self):
        if self.state["daily_count"] >= self.daily_limit:
            raise RuntimeError(f"Daily limit reached: {self.daily_limit}")
        last = self.state.get("last_request_epoch")
        if last:
            desired = random.uniform(self.minimum_interval, self.maximum_interval)
            elapsed = time.time() - float(last)
            if elapsed < desired:
                interruptible_sleep(desired - elapsed, "Rate limit")
        if (
            self.batch_size
            and self.state["daily_count"]
            and self.state["daily_count"] % self.batch_size == 0
        ):
            interruptible_sleep(self.batch_break_seconds, "Scheduled batch break")

    def record_request(self):
        self.state["daily_count"] += 1
        self.state["last_request_epoch"] = time.time()
        self.save()


def block_reason(body_text: str, response_statuses: Iterable[int]) -> str | None:
    blocked_status = next(
        (status for status in response_statuses if status in {403, 429}),
        None,
    )
    if blocked_status:
        return f"HTTP {blocked_status} observed"
    lowered = body_text.lower()
    for pattern in BLOCK_TEXT_PATTERNS:
        if pattern in lowered:
            return f"Block/CAPTCHA text observed: {pattern}"
    return None


def formula_results_json(results: Mapping[str, Iterable[FormulaCandidate]]):
    return {
        formula: [
            {
                "power_d": candidate.power_d,
                "predicted_refraction_d": candidate.predicted_refraction_d,
            }
            for candidate in candidates
        ]
        for formula, candidates in results.items()
    }


def latest_checkpoint_records(path: Path):
    latest = {}
    if not path.exists():
        raise FileNotFoundError(f"Checkpoint not found: {path}")
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        record = json.loads(line)
        public_id = record.get("public_id")
        if public_id:
            latest[public_id] = record
    return latest


def flattened_result(record):
    summary = record.get("summary") or {}
    refractions = summary.get("common_power_refractions_d") or {}
    flattened = {
        "ESCRS_Public_ID": record.get("public_id"),
        "ESCRS_Status": record.get("status"),
        "ZCB00_Barrett_Anchor_Power_D": summary.get("anchor_power_d"),
        "ZCB00_Barrett_Anchor_Pred_SE_D": summary.get(
            "anchor_predicted_refraction_d"
        ),
        "ZCB00_Formula_Count_at_Anchor": summary.get("formula_count_at_anchor"),
        "ZCB00_Formula_Spread_D": summary.get("formula_spread_d"),
    }
    column_names = {
        "Barrett Universal II": "Pred_SE_Barrett_ZCB00_Anchor_D",
        "Cooke K6": "Pred_SE_Cooke_K6_ZCB00_Anchor_D",
        "EVO 2.0": "Pred_SE_EVO_ZCB00_Anchor_D",
        "Hill-RBF 3.0": "Pred_SE_Hill_RBF_ZCB00_Anchor_D",
        "Hoffer QST": "Pred_SE_Hoffer_QST_ZCB00_Anchor_D",
        "Kane": "Pred_SE_Kane_ZCB00_Anchor_D",
        "PEARL-DGS": "Pred_SE_PEARL_DGS_ZCB00_Anchor_D",
    }
    for formula, column in column_names.items():
        flattened[column] = refractions.get(formula)
    return flattened


def export_results(args) -> int:
    checkpoint_path = Path(args.checkpoint)
    records = latest_checkpoint_records(checkpoint_path)
    rows = [
        {
            "Source_Row": record.get("source_row"),
            "Source_Eye": record.get("source_eye"),
            "Source_Cohort": (record.get("research_flags") or {}).get(
                "source_cohort"
            ),
            "Prior_Refractive_Surgery_Status": (
                record.get("research_flags") or {}
            ).get("prior_refractive_surgery_status"),
            **flattened_result(record),
        }
        for record in records.values()
        if record.get("status") in {"complete", "partial", "not_evaluable"}
    ]
    if not rows:
        raise ValueError("Checkpoint contains no exportable result rows.")

    summary_path = Path(args.summary_csv)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    with summary_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote anonymized result summary: {summary_path}")

    if args.source_workbook or args.output_workbook:
        if not args.source_workbook or not args.output_workbook:
            raise ValueError(
                "--source-workbook and --output-workbook must be supplied together."
            )
        source_path = Path(args.source_workbook).resolve()
        output_path = Path(args.output_workbook).resolve()
        if source_path == output_path:
            raise ValueError("The output workbook must not overwrite the source workbook.")
        if output_path.exists() and not args.overwrite:
            raise FileExistsError(
                f"Output workbook already exists: {output_path}. "
                "Pass --overwrite to replace only the generated copy."
            )
        workbook = load_workbook(source_path)
        worksheet = (
            workbook[args.sheet]
            if args.sheet
            else choose_sheet(
                workbook,
                None,
                {"Pat_ID", "Acquisition_Date", "Eye_Side", "AL"},
            )
        )
        result_columns = list(flattened_result({}))
        existing_headers = {
            str(cell.value).strip(): cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }
        for column_name in result_columns:
            if column_name not in existing_headers:
                column = worksheet.max_column + 1
                worksheet.cell(row=1, column=column, value=column_name)
                existing_headers[column_name] = column
        by_row = {
            int(row["Source_Row"]): row
            for row in rows
            if row.get("Source_Row") is not None
        }
        for source_row, row in by_row.items():
            for column_name in result_columns:
                worksheet.cell(
                    row=source_row,
                    column=existing_headers[column_name],
                    value=row.get(column_name),
                )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        workbook.close()
        print(f"Wrote workbook copy with ESCRS results: {output_path}")
    return 0


def run_live(args) -> int:
    from playwright.sync_api import sync_playwright

    profile = read_json(Path(args.profile))
    if not profile.get("verified"):
        raise RuntimeError(
            "The UI profile is not verified. Run inspect-ui and validate selectors first."
        )
    if args.confirm_live != LIVE_CONFIRMATION:
        raise RuntimeError(
            "Live execution requires --confirm-live " + LIVE_CONFIRMATION
        )
    queue = read_json(Path(args.queue))
    if queue.get("contains_patient_identifiers") is not False:
        raise RuntimeError("Queue anonymization metadata is missing or invalid.")
    if not args.allow_unverified_history:
        unverified = [
            item
            for item in queue["cases"]
            if item["research_flags"]["prior_refractive_surgery_status"] == "unverified"
        ]
        if unverified:
            raise RuntimeError(
                f"{len(unverified)} cases have unverified refractive-surgery history. "
                "Review them or pass --allow-unverified-history for an exploratory pilot."
            )

    output_dir = Path(args.output_dir)
    checkpoint_path = output_dir / "checkpoint.jsonl"
    snapshot_dir = output_dir / "snapshots"
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    done = completed_ids(checkpoint_path)
    limiter = RateLimiter(
        output_dir / "rate_state.json",
        args.minimum_interval,
        args.maximum_interval,
        args.daily_limit,
        args.batch_size,
        args.batch_break_seconds,
    )

    cases = [
        item for item in queue["cases"] if item["case"]["public_id"] not in done
    ][: args.limit]
    if not cases:
        print("No pending cases.")
        return 0

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            channel=args.browser_channel,
            headless=False,
            slow_mo=args.slow_mo,
        )
        context = browser.new_context(
            locale="en-US",
            viewport={"width": 1500, "height": 1100},
        )
        page = context.new_page()
        response_statuses: list[int] = []
        page.on("response", lambda response: response_statuses.append(response.status))
        page.goto(profile.get("url", DEFAULT_URL), wait_until="domcontentloaded")
        page.wait_for_timeout(args.settle_seconds * 1000)
        try:
            ensure_calculator_ready(
                page,
                profile,
                args.consent_wait_seconds,
            )
            validate_page_profile(page, profile)
        except Exception as exc:
            save_error_snapshot(
                page,
                snapshot_dir,
                cases[0]["case"]["public_id"] + "_startup",
                str(exc),
                response_statuses,
            )
            browser.close()
            raise

        for item in cases:
            case = item["case"]
            limiter.before_case()
            response_statuses.clear()
            started_at = datetime.now().astimezone().isoformat()
            checkpoint_written = False
            try:
                if profile.get("reset_button_names"):
                    try:
                        click_named_button(page, profile["reset_button_names"])
                        page.wait_for_timeout(1000)
                    except RuntimeError:
                        page.goto(
                            profile.get("url", DEFAULT_URL),
                            wait_until="domcontentloaded",
                        )
                        page.wait_for_timeout(args.settle_seconds * 1000)
                        ensure_calculator_ready(
                            page,
                            profile,
                            args.consent_wait_seconds,
                        )
                        validate_page_profile(page, profile)
                fill_case(page, case, profile)
                validate_iol_constants(page, profile)
                click_named_button(page, profile["calculate_button_names"])
                limiter.record_request()
                body_text, formula_results = wait_for_formula_results(
                    page,
                    profile,
                    args.result_wait_seconds,
                )
                reason = block_reason(body_text, response_statuses)
                if reason:
                    append_jsonl(
                        checkpoint_path,
                        {
                            "public_id": case["public_id"],
                            "status": "halted",
                            "reason": reason,
                            "started_at": started_at,
                            "finished_at": datetime.now().astimezone().isoformat(),
                        },
                    )
                    checkpoint_written = True
                    raise RuntimeError(reason)

                summary = summarize_formula_results(formula_results)
                snapshot_base = snapshot_dir / case["public_id"]
                snapshot_base.with_suffix(".txt").write_text(
                    body_text,
                    encoding="utf-8",
                )
                page.screenshot(
                    path=str(snapshot_base.with_suffix(".png")),
                    full_page=True,
                )
                record = {
                    "public_id": case["public_id"],
                    "source_row": case["source_row"],
                    "source_eye": case["source_eye"],
                    "research_flags": item["research_flags"],
                    "status": summary["status"],
                    "started_at": started_at,
                    "finished_at": datetime.now().astimezone().isoformat(),
                    "calculator_url": page.url,
                    "formula_results": formula_results_json(formula_results),
                    "summary": summary,
                }
                append_jsonl(checkpoint_path, record)
                checkpoint_written = True
                print(
                    f"{case['public_id']}: {summary['status']} "
                    f"({summary.get('formula_count_at_anchor', 0)} formulas)"
                )
                if summary["status"] != "complete" and args.stop_on_partial:
                    raise RuntimeError(
                        "Unexpected partial result; stopped for selector review."
                    )
            except Exception as exc:
                save_error_snapshot(
                    page,
                    snapshot_dir,
                    case["public_id"],
                    str(exc),
                    response_statuses,
                )
                if not checkpoint_written:
                    append_jsonl(
                        checkpoint_path,
                        {
                            "public_id": case["public_id"],
                            "status": "error",
                            "reason": str(exc),
                            "started_at": started_at,
                            "finished_at": datetime.now().astimezone().isoformat(),
                        },
                    )
                browser.close()
                raise
        browser.close()
    return 0


def build_parser():
    parser = argparse.ArgumentParser(
        description="Conservative ESCRS public-UI formula-spread pilot."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    prepare = subparsers.add_parser(
        "prepare",
        help="Create an anonymized, validated pilot queue from a local workbook.",
    )
    prepare.add_argument("--input", required=True)
    prepare.add_argument("--sheet")
    prepare.add_argument(
        "--gender-workbook",
        default="IOLMaster700_corrected_3.xlsx",
    )
    prepare.add_argument("--gender-sheet")
    prepare.add_argument("--limit", type=int, default=10)
    prepare.add_argument(
        "--id-key",
        default=str(DEFAULT_KEY_PATH),
    )
    prepare.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_DIR / "pilot_queue.json"),
    )
    prepare.set_defaults(func=prepare_queue)

    inspect = subparsers.add_parser(
        "inspect-ui",
        help="Open the public form and save a control manifest without entering data.",
    )
    inspect.add_argument(
        "--profile",
        default="automation/escrs_ui_profile.example.json",
    )
    inspect.add_argument("--browser-channel", default="msedge")
    inspect.add_argument("--settle-seconds", type=int, default=8)
    inspect.add_argument("--consent-wait-seconds", type=float, default=45.0)
    inspect.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR / "ui_inspection"))
    inspect.add_argument("--keep-open", action="store_true")
    inspect.add_argument(
        "--probe-lens-options",
        action="store_true",
        help=(
            "Open Gender/Manufacturer/IOL dropdowns and save their public option "
            "texts without entering patient or biometry data."
        ),
    )
    inspect.set_defaults(func=inspect_ui)

    live = subparsers.add_parser(
        "run",
        help="Run a verified, visible, sequential 10-eye pilot.",
    )
    live.add_argument(
        "--queue",
        default=str(DEFAULT_OUTPUT_DIR / "pilot_queue.json"),
    )
    live.add_argument(
        "--profile",
        default="automation/escrs_ui_profile.json",
    )
    live.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR / "pilot_run"))
    live.add_argument("--browser-channel", default="msedge")
    live.add_argument("--limit", type=int, default=10)
    live.add_argument("--minimum-interval", type=float, default=45.0)
    live.add_argument("--maximum-interval", type=float, default=75.0)
    live.add_argument("--daily-limit", type=int, default=100)
    live.add_argument("--batch-size", type=int, default=25)
    live.add_argument("--batch-break-seconds", type=float, default=600.0)
    live.add_argument("--settle-seconds", type=int, default=8)
    live.add_argument("--consent-wait-seconds", type=float, default=45.0)
    live.add_argument("--result-wait-seconds", type=int, default=90)
    live.add_argument("--slow-mo", type=int, default=150)
    live.add_argument("--allow-unverified-history", action="store_true")
    live.add_argument(
        "--stop-on-partial",
        action=argparse.BooleanOptionalAction,
        default=True,
    )
    live.add_argument("--confirm-live", default="")
    live.set_defaults(func=run_live)

    export = subparsers.add_parser(
        "export",
        help="Export checkpoint results to an anonymized CSV and optional workbook copy.",
    )
    export.add_argument(
        "--checkpoint",
        default=str(DEFAULT_OUTPUT_DIR / "pilot_run/checkpoint.jsonl"),
    )
    export.add_argument(
        "--summary-csv",
        default=str(DEFAULT_OUTPUT_DIR / "pilot_run/results_summary.csv"),
    )
    export.add_argument("--source-workbook")
    export.add_argument("--sheet")
    export.add_argument("--output-workbook")
    export.add_argument("--overwrite", action="store_true")
    export.set_defaults(func=export_results)
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()
    try:
        return args.func(args)
    except KeyboardInterrupt:
        print("Stopped by user.", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
