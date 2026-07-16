import json
import tempfile
import unittest
from argparse import Namespace
from pathlib import Path

from openpyxl import Workbook, load_workbook

from automation.escrs_formula_pilot import export_results
from automation.escrs_pilot_core import (
    FormulaCandidate,
    assert_anonymized_payload,
    formula_spread,
    make_public_id,
    normalize_cct_um,
    normalize_gender,
    prepare_case,
    radius_mm_to_diopters,
    select_anchor_candidate,
    select_evenly_across_al,
    summarize_formula_results,
)


class EsCRSFormulaPilotTests(unittest.TestCase):
    def test_radius_conversion_and_cct_units(self):
        self.assertAlmostEqual(radius_mm_to_diopters(7.5), 45.0)
        self.assertEqual(normalize_cct_um(0.55), 550.0)
        self.assertEqual(normalize_cct_um(550), 550)

    def test_gender_normalization(self):
        self.assertEqual(normalize_gender("M"), "Male")
        self.assertEqual(normalize_gender("여성"), "Female")
        self.assertIsNone(normalize_gender(""))

    def test_public_id_is_stable_and_does_not_contain_source_identifier(self):
        secret = b"0123456789abcdef0123456789abcdef"
        first = make_public_id(secret, "patient-123", "2026-01-02", "OD")
        second = make_public_id(secret, "patient-123", "2026-01-02", "R")
        self.assertEqual(first, second)
        self.assertNotIn("patient", first.lower())
        self.assertRegex(first, r"^ESCRS-[A-F0-9]{12}$")

    def test_prepare_case_emits_anonymized_browser_payload(self):
        row = {
            "Pat_ID": "secret-patient",
            "Last_Name": "Sensitive",
            "First_Name": "Name",
            "DOB": "1970-01-01",
            "Acquisition_Date": "2026-01-02",
            "Eye_Side": "OD",
            "AL": 24.0,
            "R1": 7.7,
            "R2": 7.5,
            "ACD": 3.2,
            "LT": 4.5,
            "CCT": 0.54,
            "W2W": 11.8,
        }
        case = prepare_case(
            row,
            source_row=12,
            gender="Female",
            id_secret=b"0123456789abcdef0123456789abcdef",
        )
        payload = case.browser_payload()
        serialized = json.dumps(payload)
        self.assertNotIn("secret-patient", serialized)
        self.assertNotIn("Sensitive", serialized)
        self.assertNotIn("1970-01-01", serialized)
        self.assertNotIn("source_row", payload)
        self.assertEqual(payload["calculator_slot"], "right")
        assert_anonymized_payload(payload)

    def test_payload_rejects_identifier_fields(self):
        with self.assertRaises(ValueError):
            assert_anonymized_payload({"patient_id": "123"})

    def test_anchor_is_inside_band_and_closest_to_minus_quarter(self):
        candidates = [
            FormulaCandidate(20.0, 0.10),
            FormulaCandidate(20.5, -0.18),
            FormulaCandidate(21.0, -0.44),
        ]
        selected = select_anchor_candidate(candidates)
        self.assertEqual(selected, FormulaCandidate(20.5, -0.18))

    def test_anchor_tie_prefers_slightly_more_myopic_refraction(self):
        candidates = [
            FormulaCandidate(20.0, -0.10),
            FormulaCandidate(20.5, -0.40),
        ]
        selected = select_anchor_candidate(candidates)
        self.assertEqual(selected, FormulaCandidate(20.5, -0.40))

    def test_anchor_rejects_power_outside_zcb00_half_diopter_grid(self):
        selected = select_anchor_candidate(
            [FormulaCandidate(20.25, -0.25)]
        )
        self.assertIsNone(selected)

    def test_formula_spread_uses_same_anchor_power(self):
        results = {
            "Barrett Universal II": [
                FormulaCandidate(20.0, 0.12),
                FormulaCandidate(20.5, -0.22),
                FormulaCandidate(21.0, -0.55),
            ],
            "Cooke K6": [
                FormulaCandidate(20.0, 0.20),
                FormulaCandidate(20.5, -0.10),
                FormulaCandidate(21.0, -0.43),
            ],
            "Kane": [
                FormulaCandidate(20.0, 0.02),
                FormulaCandidate(20.5, -0.34),
                FormulaCandidate(21.0, -0.70),
            ],
        }
        summary = summarize_formula_results(results)
        self.assertEqual(summary["anchor_power_d"], 20.5)
        self.assertAlmostEqual(summary["formula_spread_d"], 0.24)
        self.assertEqual(summary["formula_count_at_anchor"], 3)

    def test_spread_requires_at_least_three_formulas(self):
        self.assertIsNone(formula_spread({"A": -0.2, "B": -0.4}))

    def test_even_sampling_covers_axial_length_range(self):
        row = {
            "Pat_ID": "p",
            "Acquisition_Date": "2026-01-02",
            "Eye_Side": "R",
            "R1": 7.7,
            "R2": 7.5,
            "ACD": 3.2,
            "LT": 4.5,
            "CCT": 0.54,
            "W2W": 11.8,
        }
        cases = []
        for index, al in enumerate((16.0, 20.0, 24.0, 28.0, 32.0)):
            item = dict(row, AL=al, Pat_ID=f"p{index}")
            cases.append(
                prepare_case(
                    item,
                    source_row=index + 2,
                    gender="M",
                    id_secret=b"0123456789abcdef0123456789abcdef",
                )
            )
        selected = select_evenly_across_al(cases, 3)
        self.assertEqual([case.al_mm for case in selected], [16.0, 24.0, 32.0])

    def test_export_writes_copy_without_overwriting_source(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "source.xlsx"
            output = root / "output.xlsx"
            summary_csv = root / "summary.csv"
            checkpoint = root / "checkpoint.jsonl"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Study_Cases"
            worksheet.append(["Pat_ID", "Acquisition_Date", "Eye_Side", "AL"])
            worksheet.append(["secret", "2026-01-02", "R", 24.0])
            workbook.save(source)
            workbook.close()

            record = {
                "public_id": "ESCRS-ABCDEF123456",
                "source_row": 2,
                "source_eye": "R",
                "research_flags": {
                    "source_cohort": "Control",
                    "prior_refractive_surgery_status": "unverified",
                },
                "status": "complete",
                "summary": {
                    "anchor_power_d": 20.5,
                    "anchor_predicted_refraction_d": -0.2,
                    "formula_count_at_anchor": 3,
                    "formula_spread_d": 0.3,
                    "common_power_refractions_d": {
                        "Barrett Universal II": -0.2,
                        "Cooke K6": -0.1,
                        "Kane": -0.4,
                    },
                },
            }
            checkpoint.write_text(
                json.dumps(record) + "\n",
                encoding="utf-8",
            )
            export_results(
                Namespace(
                    checkpoint=str(checkpoint),
                    summary_csv=str(summary_csv),
                    source_workbook=str(source),
                    output_workbook=str(output),
                    sheet="Study_Cases",
                    overwrite=False,
                )
            )

            original = load_workbook(source, read_only=True, data_only=True)
            self.assertEqual(original["Study_Cases"].max_column, 4)
            original.close()
            generated = load_workbook(output, read_only=True, data_only=True)
            headers = [
                cell.value
                for cell in next(
                    generated["Study_Cases"].iter_rows(min_row=1, max_row=1)
                )
            ]
            spread_column = headers.index("ZCB00_Formula_Spread_D") + 1
            self.assertEqual(
                generated["Study_Cases"].cell(row=2, column=spread_column).value,
                0.3,
            )
            generated.close()
            self.assertTrue(summary_csv.exists())


if __name__ == "__main__":
    unittest.main()
